###################################################################################
#	Marvell GPL License
#	
#	If you received this File from Marvell, you may opt to use, redistribute and/or
#	modify this File in accordance with the terms and conditions of the General
#	Public License Version 2, June 1991 (the "GPL License"), a copy of which is
#	available along with the File in the license.txt file or by writing to the Free
#	Software Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA 02111-1307 or
#	on the worldwide web at http://www.gnu.org/licenses/gpl.txt.
#	
#	THE FILE IS DISTRIBUTED AS-IS, WITHOUT WARRANTY OF ANY KIND, AND THE IMPLIED
#	WARRANTIES OF MERCHANTABILITY OR FITNESS FOR A PARTICULAR PURPOSE ARE EXPRESSLY
#	DISCLAIMED.  The GPL License provides additional details about this warranty
#	disclaimer.
###################################################################################

from __future__ import print_function
from builtins import str
from builtins import object
__author__ = "Alex Zemtzov"

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
import os.path
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

class OpenpyXlWrite(object):

    def __init__(self):
        self._wb = Workbook()
        self._ws = self._wb.active
        self._FileName = ""
        self._m_CurrWorkSheet = None

    def SetActiveSheet(self, i_NewActiveSheetName):
        """
        The method is setting the active sheet as i_NewActiveSheetName.
        If the sheet exist then we append to it, if the sheet not exists, then we create it.
        as well the 'SHEET1' is deleted.
        NO SHEETS WITH 'SHEET' can be in file
        :param i_NewActiveSheetName:
        :return:bool value if the sheet exist or new one was created
        """
        isSheetExist = False
        numOfActiveSheet = -1
        allSheet = self._wb.sheetnames

        for numOfSheet, sheet in enumerate(allSheet):
            if i_NewActiveSheetName.lower() == str(sheet).lower():
                isSheetExist = True
                numOfActiveSheet = numOfSheet
        if not isSheetExist:
            self._wb.create_sheet(i_NewActiveSheetName, 0)
            self._wb.active = 0
        else:
            self._wb.active = numOfActiveSheet

        self._ws = self._wb.active
        allSheet = self._wb.sheetnames
        for numOfSheet, sheet in enumerate(allSheet):
            if 'sheet' in str(sheet).lower():
                sheetToRemove = self._wb[sheet]
                self._wb.remove(sheetToRemove)

        return isSheetExist

    def SetFileName(self, i_FileName):
        """
        The method is setting the active file name, if the file exist then we append
        to the existing file to random sheet
        :param i_FileName:
        :return:
        """
        isFileExist = os.path.exists(i_FileName)
        self._wb = Workbook()
        if not isFileExist:
            self._wb.save(i_FileName)

        self._FileName = i_FileName
        self._wb = load_workbook(i_FileName)

    def GetActiveFileName(self):
        """
        The method is returning the active file name
        :return:
        """
        return self._FileName

    def GetActiveSheetName(self):
        """
        The method returns the active sheet name
        :return:
        """
        return self._wb.active.title

    def WriteToCell(self, i_CellLocation, i_CellValue):
        """
        The method is writing to the active file, active sheet,
        to givven cell the value that was passed
        :param i_CellNumeber:
        :param i_CellValue:
        :return:
        """
        try:
            dataToWrite = int(i_CellValue)
        except ValueError:
            try:
                dataToWrite = float(i_CellValue)
            except:
                dataToWrite = str(i_CellValue)
        finally:
            self._ws[i_CellLocation] = dataToWrite

    def GetValueFromCell(self, i_CellLocation):
        """
        The method returns a value that is written in the givven cell
        :param i_CellLocation:
        :return:
        """
        return self._ws[i_CellLocation]

    def ApplyFormat(self, i_CellNumber, i_FontFormatDict = None, i_CellFormatDict = None):
        """
        The method is setting the givven format to the givven cell,
        please notice that this method can be modified in the future while new
        format params can be added
        :param i_CellNuber:
        :param i_FormatData:
        :return:
        """
        try:
            self._ws[i_CellNumber].font = Font(sz=i_FontFormatDict['size'], color=i_FontFormatDict['color'])
            self._ws[i_CellNumber].fill = PatternFill(fill_type=i_CellFormatDict['fill_type'],
                                                      start_color=i_CellFormatDict['start_color'],
                                                      end_color=i_CellFormatDict['end_color'])
        except Exception as ex:
            print(ex.args)
            print(str(ex))

    def ConditionalFormat(self, i_Formula, i_CellLocation):
        """
        The method is suporting the conditional formating,
        ONLY BOOLEAN,
        GETS A formula, and writes PASS / FAIL and makes the cell green or red
        :param i_Formula:
        :param i_CellLocation:
        :return:
        """

        redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
        greenFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

        self._ws[i_CellLocation].value = '=IF({},"Pass","Fail")'.format(i_Formula)
        formulaString = 'NOT(ISERROR(SEARCH("Pass",{})))'.format(i_CellLocation)

        self._ws.conditional_formatting.add(i_CellLocation, FormulaRule(formula=[formulaString], stopIfTrue=True,
                                                                 fill=greenFill))
        formulaString = 'NOT(ISERROR(SEARCH("Fail",{})))'.format(i_CellLocation)

        self._ws.conditional_formatting.add(i_CellLocation,
                                      FormulaRule(formula=[formulaString], stopIfTrue=True, fill=redFill))

    def Save(self):
        """
        The methos is saving the current work in the file that was givven
        :return:
        """
        try:
            self._wb.save(self._FileName)

        except Exception as ex:
            print(ex.args)
            print(str(ex))