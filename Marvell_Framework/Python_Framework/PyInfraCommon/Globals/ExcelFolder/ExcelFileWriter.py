###################################################################################
#	Marvell GPL License
#	
#	If you received this File from Marvell, you may opt to use, redistribute and/or
#	modify this File in accordance with the terms and conditions of the General
#	Public License Version 2, June 1991 (the "GPL License"), a copy of which is
#	available along with the File in the license.txt file or by writing to the Free
#	Software Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA 02111-1307 or
#	on the worldwide web at http://www.gnu.org/licenses/gpl.txt.
#	
#	THE FILE IS DISTRIBUTED AS-IS, WITHOUT WARRANTY OF ANY KIND, AND THE IMPLIED
#	WARRANTIES OF MERCHANTABILITY OR FITNESS FOR A PARTICULAR PURPOSE ARE EXPRESSLY
#	DISCLAIMED.  The GPL License provides additional details about this warranty
#	disclaimer.
###################################################################################

from __future__ import print_function
from __future__ import absolute_import
from builtins import chr
from builtins import str
from builtins import range
from builtins import object
__author__ = "Alex Zemtzov, Erez Ashkenazi"

from . import OpenpyXlWrite
from openpyxl.utils.cell import get_column_letter, column_index_from_string

class CellFormating(object):
    """
    This class is sub class for ExcelFileWriter
    """

    def __init__(self, *i_DefaultParams):
        """
        It is possible to set a new default formatig
        :param i_DefaultParams:
        cellColor
        fontSize
        if you want to add more stuff please add set method and add another case to 'SetDefaultFormat' method
        """
        self._m_Colors = {'black':"000000", 'blue':"0000FF",'green' : "00E75C", 'yellow' : "FFF711",
                            'red' : 'fb7070',"aquamarine":'7FFFD4','jellyfish' : '46C7C7','cream' : 'ffffcc',
                            'white' : 'FFFFFF', 'pass': '00FF00', 'fail':'FF0000', 'na':"0000FF"}

        self._m_DefaultParams = ['cellColor=white', 'fontSize=11']
        self.SetDefaultFormat()
        if(len(i_DefaultParams[0]) > 0):
            self._m_DefaultParams = ""
            for subFormat in i_DefaultParams[0]:
                self._m_DefaultParams += " " + subFormat
            self.SetDefaultFormat()

    def GetAllPossibleColors(self):
        return list(self._m_Colors.keys())

    def SetNewCellColor(self, i_Color):
        """
        The method is setting the color of the cell,
        if the color not in list of m_Colors,
        the color that will be is default - White
        :param i_Color:
        :return:
        """
        colorToSet = ''
        allColors = self.GetAllPossibleColors()
        if str(i_Color).lower() not in allColors:
            colorToSet = 'white'
        else:
            colorToSet = str(i_Color).lower()

        self._m_CellColor = colorToSet

    def SetNewFontsize(self, i_FontSize):
        """
        The method is seting the font size to new one,
        if the font size is givven not properly then the size will.
        The Max size is 32
        be as default: 14
        :param i_FontSize:
        :return:
        """
        newSize = 0
        try:
            newSize = int(i_FontSize)
            if newSize > 48:
                newSize = 32
            if newSize < 0:
                raise ValueError

        except (ValueError):
            newSize = 14

        self._m_CellFontSize = newSize

    def GetCellColor(self):
        """
        The method is returnig the HEX value of the color
        :return:
        """
        return self._m_Colors[self._m_CellColor]

    def GetCellFontSize(self):
        return self._m_CellFontSize

    def GetCellFontColor(self):
        return self._m_Colors['black']

    def GetSpecificColor(self, i_NameOfColor):
        """
        The method is returnig the hex value of specific key,
        if the key is not found the answer will be white
        :param i_NameOfColor:
        :return:
        """
        answer = self._m_Colors['black']
        allColors = self.GetAllPossibleColors()
        if str(i_NameOfColor).lower() in allColors:
            answer =  self._m_Colors[str(i_NameOfColor).lower()]

        return answer

    def SetDefaultFormat(self):
        """
        The method is setting the format for default settings that were assigned when
        the class was created
        :return:
        """
        if type(self._m_DefaultParams) == str:
            dataAsList = str(self._m_DefaultParams).split()
        else:
            dataAsList = self._m_DefaultParams

        for param in dataAsList:
            paramWithValue = param.split('=')
            if str(paramWithValue[0]).lower() == 'cellcolor':
                self.SetNewCellColor(str(paramWithValue[1]))
            elif str(paramWithValue[0]).lower() == 'fontsize':
                self.SetNewFontsize(str(paramWithValue[1]))

    def AddNewColor(self, i_ColorName, i_ColorValue):
        """
        This method can add new color to the existing colors dictionary
        :param i_ColorName: should be lower case
        :param i_ColorValue: should be in HEX format - example '7FFFD4'
        :return:
        """
        import re

        match = re.search(r'(?:[0-9a-fA-F]{6})', i_ColorValue)
        try:
            if match:
                if(i_ColorName not in list(self._m_Colors.keys())):
                    self._m_Colors[i_ColorName] = i_ColorValue;
                else:
                    raise Exception('The Name Of The Color Already Exist')
            else:
                raise Exception('The Hex Value Of The Color is not written in proper way, please see HLP')
        except Exception as error:
            print(error.args)

class LocationOnSheet(object):
    """
    The Class is made for Excel module,
    some methods may be added in the future.
    The Class is initiliazed with location of 'A1'.
    The Class is Not Supporting Yet two letters Collumns

    """

    def __init__(self, i_NewLocation = None):
        self._m_CurrRow = 1
        self._m_CurrColl = 1  #Todo: as a WA usage of get_column_letter, column_index_from_string was implemented.
                              # need to rewrite the section of nextColumn/line
        self._m_LastRow = 1
        self._m_LastColl = 1
        if i_NewLocation != None:
            self.SetNewLocation(i_NewLocation)

    def SetCurrRow(self, i_NewRowLoc):
        """
        The method is set method for self._m_CurrRow  var
        :param i_NewRowLoc: must be  positive int
        :return:
        """
        isInputCorrect = False
        try:
            val ,isInputCorrect = self._intTryParse(i_NewRowLoc)
            if isInputCorrect:
                self._m_LastRow = val
                self._m_CurrRow = val
            else:
                raise ValueError

        except ValueError:
            print("You Entered a bad value for new Location Of row your input {userInput}".format(userInput = i_NewRowLoc))

    def SetCurrColl(self, i_NewCollLoc):
        """
        The method is set method for self._m_CurrColl  var
        :param i_NewRowLoc: must be int
        :return:
        """
        isInputCorrect = False
        try:
            value, isInputCorrect = self._CheckCollInput(i_NewCollLoc)
            if isInputCorrect:
                self._m_LastColl = self._m_CurrColl
                self._m_CurrColl = column_index_from_string(value)
            else:
                raise ValueError

        except ValueError:
            print("You Entered a bad value for new Location Of Coll your input {userInput}".format(userInput = i_NewCollLoc))

    def SetNewLocation(self, i_NewLocation):
        """
        The new location should be a string as following: 'B6' or int describing the column number(index)
        If there is any syntax error  the method will raise exception
        :param i_NewLocation: param for new location of row and coll
        :return:
        """
        collNewLoc = ""
        rowNewLoc = ""
        if type(i_NewLocation) == int:
            i_NewLocation = get_column_letter(i_NewLocation)

        try:
            collNewLoc, rowNewLoc = self._ParseUserLocation(i_NewLocation)
            if type(collNewLoc) is str and type(rowNewLoc) is int and len(collNewLoc) > 0:
                self._m_LastRow = self._m_CurrRow
                self._m_LastColl = self._m_CurrColl
                self.SetCurrColl(collNewLoc)
                self.SetCurrRow(rowNewLoc)
            else:
                raise ValueError
        except ValueError:
            print("You Entered a Bad Value for new location, your value is {userInput}".format(userInput = i_NewLocation))

    def _intTryParse(self, i_ValueToCheck):
        """
        The method is checkinf if input is intreger and greater then 0
        :return:
        """
        try:
             int(i_ValueToCheck)
             if int(i_ValueToCheck) < 0:
                 raise  ValueError
             else:
                 return int(i_ValueToCheck), True
        except ValueError:
            return i_ValueToCheck, False

    def _CheckCollInput(self, i_NewCollLoc):
        """
        The Method is checkig if the input has only capital letters i_NewCollLoc
        :param i_NewCollLoc:
        :return:
        """
        isInputCorrect = True
        try:
            if len(i_NewCollLoc) < 1:
                isInputCorrect = False
                raise ValueError
            elif len(i_NewCollLoc) == 1:
                    if (ord(i_NewCollLoc) < ord('A') or ord(i_NewCollLoc) > ord('Z')):
                        isInputCorrect = False
            else:
                for char in i_NewCollLoc:
                    if ord(char) < ord('A') or ord(char) > ord('Z'):
                        isInputCorrect = False
                        raise ValueError
        except ValueError:
            print("You Entered a bad value for new Location Of coll. " \
                  "Your input: '{userInput}'".format(userInput = i_NewCollLoc))
        return i_NewCollLoc, isInputCorrect

    def _ParseUserLocation(self, i_NewLocation):
        """
        The method is only parsing the input not checking if there was an error
        in the input.
        Self Correction is done if the letters (New Coll Location) were givven as lower Case.
        :param i_NewLocation:
        :return:
        """
        isOnlyNumbersLeft = False
        i_NewLocation = str(i_NewLocation).upper()
        collNewLoc = ""
        rowNewLoc = ""
        for char in i_NewLocation:
            if ord(char) >= ord('A') and ord(char) <= ord('Z') and isOnlyNumbersLeft == False:
                collNewLoc += char
            else:
                isOnlyNumbersLeft = True
                rowNewLoc += char

        rowNewLoc, isParseSucceed = self._intTryParse(rowNewLoc)
        return collNewLoc, rowNewLoc

    def GetCurrLocaion(self, i_ReturnAsDicOrStr = 'str'):
        """

        :param i_ReturnAsDicOrStr: if you want a return value as dictionary - {coll:'c',row:'5'}
         then send 'dic' param to file
        :return:
        """
        currLocation = {}
        currLocation['Row'] = self._m_CurrRow
        currLocation['Coll'] = get_column_letter(self._m_CurrColl)
        if i_ReturnAsDicOrStr == 'dic':
            return currLocation
        else:
            return  str(currLocation['Coll']) + str(self._m_CurrRow)

    def GoToNextLine(self):
        """
        The method is setting the curr location as
        collumn A
        row - last row + 1
        please notice that we do save last location before moving to next line
        :return:
        """
        currLocation = self.GetCurrLocaion('dic')
        currRow = int(currLocation['Row'])
        currColl = currLocation['Coll']
        self._m_LastRow = currRow
        self._m_LastColl = currColl
        currRow += 1
        currColl = 'A'
        newLocation = str(currColl) + str(currRow)
        self.SetNewLocation(newLocation)

    def GoToNextColl(self):
        """
        The method is incrementig the curr coll
        :return:
        """
        currColl = self.GetCurrLocaion('dic')['Coll']
        self._m_LastColl = currColl
        self._m_CurrColl += 1
        newCurrColl = get_column_letter(self._m_CurrColl)
        self.SetCurrColl(newCurrColl)

    def _IncrementCollWithMultiLetter(self, io_CurrColl):
        """
        if the collums have more then one letter, then this incrimentin will work properly
        :param io_CurrColl:
        :return:
        """
        collAsList = []
        newColl = []
        isMoreChangedToDo = True
        for char in io_CurrColl:
            collAsList.append(char)
        collAsList = reversed(collAsList)
        for letter in collAsList:
            if ord(letter) < ord('Z') and isMoreChangedToDo == True:
                newLetter = chr(ord(letter) + 1)
                isMoreChangedToDo = False
                newColl.append(newLetter)
            elif ord(letter) == ord('Z'):
                newLetter= 'AA'
                newColl.append(newLetter)
            else:
                newColl.append(letter)

        newColl = reversed(newColl)
        newCollAsString = ''.join(str(x) for x in newColl)
        return newCollAsString

    def GetLastLocation(self, i_ReturnAsDicOrStr = 'str'):
        """

        :param i_ReturnAsDicOrStr: if you want a return value as dictionary - {coll:'c',row:'5'}
         then send 'dic' param to file
        :return:
        """
        lastLocation = {}
        lastLocation['Row'] = self._m_LastRow
        lastLocation['Coll'] = get_column_letter(self._m_LastColl)
        if i_ReturnAsDicOrStr == 'dic':
            return lastLocation
        else:
            return str(lastLocation['Coll']) + str(self._m_LastRow)

class ExcelFileWriter(object):
    """
    The class is not working with dictionaries,
    only strings that the separator is whitespace or comma (can override the 'ParseInputParams' method to add another separator)
    NOTICE:
    before using the methods of writing ALWAYS first set the file name and the work sheet ('SetFileNameToSaveResult', 'SetActiveWorkSheet')
    """

    def __init__(self, *i_FormatDefault):
        """
        It is possible to set a default style for writing the excel file
        by passing the default vaules here.
        please see 'CellFormating' class to get the names of default values
        :param i_SheetName:
        :param i_FileName:
        :param i_FormatDefault:
        """
        self._XlWrapper = OpenpyXlWrite.OpenpyXlWrite()
        self._CurrFileName =  None
        self._m_Locator =  LocationOnSheet()
        self._m_Formater = CellFormating(i_FormatDefault)
        self._m_NumOfCellsThatWereWrittenLastTime = 0
        self._m_LastStartingLocationOfWritting = 'A1'
        self._m_LastLocationOnSheet = {}
        self._SeparatorBetweenFormatAndData = "~"
        self._CellWithFormating = []


    def SetFileNameToSaveResult(self, i_FileName):
        """
        The method is checking if the file exists or not.
        If do exist then the file is reopened and the data will
        be appended. If the file is not exist the file will be created
        :param i_FileName:
        :return:
        """
        self._XlWrapper.SetFileName(str(i_FileName))

    def SetActiveWorkSheet(self, i_WorkSheetName):
        """
        The method is setting an active worksheet
        with the givven name.
        If the sheet exist then we append the data to it, as well we set the next place to write the last place were we been las time
        otherwise the sheet will be created.
        :param i_WorkSheetName:
        :return:
        """
        isSheetExist = False
        startLocation = 'A1'
        isSheetExist = self._XlWrapper.SetActiveSheet(i_WorkSheetName)
        if isSheetExist:
            currActiveSheet = self._XlWrapper.GetActiveSheetName()
            if self._m_LastLocationOnSheet: # if the sheet exist, and on current run we havent wrote to it, the start location will be A1
                startLocation = self._m_LastLocationOnSheet[currActiveSheet]

        self._m_Locator.SetNewLocation(startLocation)

    def WriteLineWithFormating(self, i_Data, i_save_To_File=False):
        """
        The method is writing the string with it formating to the Excel sheet that was
        has been set as curent active sheet
        :param i_Data: can be wirtten as following:
        "startLocation=C3 cellColor=yellow fontSize=12  <text whitespace separator>   cellColor=red fontSize=16 <text whitespace separator>"
        ['startLocation=C3', 'cellColor=yellow', 'fontSize=12'  <text comma separator>   'cellColor=red', 'fontSize=16' <text comma separator>]
        :param
        :return:
        """
        dataWithItFormat = self._ParseFormatingAndData(i_Data)  # list as ['format1...|data1,','format2..|data2,']
        for subData in dataWithItFormat:
                formatData = str(subData).split(self._SeparatorBetweenFormatAndData)  # [0] - format, [1] - data
                self._SetFormat(formatData)
                self._WriteData(formatData[1])
                self._SaveNextLocationPerSheet()
                self._ApplyFormat()
        self._XlWrapper.Save() if i_save_To_File else None

    def GotoNextLineOrNewLocation(self, i_StartLocation = None):
        """
        The method is setting the the writer to next line,
        to coll A.
        If you want you can give a uniqe location to set the location
        :param i_StartLocation:
        :return:
        """
        newLocation = ""
        if i_StartLocation != None:
            self._m_Locator.SetNewLocation(i_StartLocation)
        else:
            self._m_Locator.GoToNextLine()

    def SetDefaultFormat(self):
        self._m_Formater.SetDefaultFormat()

    def _ParseInputParams(self, i_Data):
        """
        The method is taking the data, and parsing it to list.
        Please Notice that only ONE param can be parsed each time.
        Don't send thigs like:
        "a1,a2 a3,a4    b1,b2  b3"
        :param i_Data: can be string with ' ' (whitespace ) or ',' (comma) as seperator,

        :return: list with same data
        """
        if type(i_Data) is list and len(i_Data) > 1:
            dataAsList = i_Data
        else:
            dataWithCommas = str(i_Data).replace(' ', ',')
            dataAsList = str(dataWithCommas).split(',')

        return dataAsList

    def _SetFormat(self, i_DataWithFormat):
        if "location" in i_DataWithFormat[0].lower():
            newStartLocation = self._GetSpecificDataFromFormat(i_DataWithFormat[0].split(','), "location")
            self._m_Locator.SetNewLocation(newStartLocation)
        if 'cellcolor' in i_DataWithFormat[0].lower():
            newColor = self._GetSpecificDataFromFormat(i_DataWithFormat[0].split(','), "cellcolor")
            self._m_Formater.SetNewCellColor(newColor)
        if 'fontsize' in i_DataWithFormat[0].lower():
            newFontSize = self._GetSpecificDataFromFormat(i_DataWithFormat[0].split(','), "fontsize")
            self._m_Formater.SetNewFontsize(newFontSize)

    def _ParseFormatingAndData(self, i_Data):
        """
        The method is parsig the input for 2 categories:
        cell formats,
        cell data.
        the input must to be as folowing:
        "startLocation=H1 cellColor=red fontSize=16 And Again We Will cellColor=yellow fontSize=10 You Will Never Die"
        while the format info can be not written at all
        :param i_Data:
        :return:
        """
        isMoreFormatingToCome = True
        data = ""
        answer = []
        formating = ""
        dataWithItFormat = ""

        dataAsList = self._ParseInputParams(i_Data)
        for param in dataAsList:
            if "=" in param and 'formula' not in param and 'FormatCellCustom'.lower() not in param.lower() and "cond_format" not in param:
                if isMoreFormatingToCome is False:
                    dataWithItFormat = formating + self._SeparatorBetweenFormatAndData + data
                    formating = ""
                    data = ""
                    answer.append(dataWithItFormat)
                    isMoreFormatingToCome = True
                formating += param + ','
            else:
                data += param + ','
                isMoreFormatingToCome = False
        dataWithItFormat = formating + self._SeparatorBetweenFormatAndData + data
        answer.append(dataWithItFormat)
        return answer

    def _GetSpecificDataFromFormat(self, i_FormatData, i_TypeOfData):
        """
        The method is getting a strig while somewhere must appear
        '***location= A1', A1- or any other location on excel sheet.
         The method is returnig the value of the location
        :param i_FormatData:
        :return:
        """
        answer = ""
        for subData in i_FormatData:
            if i_TypeOfData in subData.lower():
                typeWithValue = str(subData).split('=')
                answer = typeWithValue[-1]
        return answer

    def _WriteData(self, i_DataToWrite):
        """
        The method is writing the data from the starting point wich  located in
        self._m_Locator.GetCurrLocaion(), after writng the value the pointer is movig to next
        cell to the right.
        As well the method is counting how many cells were written
        :param i_DataToWrite:
        :return:
        """
        dataToWrite = i_DataToWrite.split(',')
        self._m_NumOfCellsThatWereWrittenLastTime = 0
        self._m_LastStartingLocationOfWritting = self._m_Locator.GetCurrLocaion('str')
        for subData in dataToWrite:
            if len(subData) >= 1:
                location = self._m_Locator.GetCurrLocaion('str')
                if "formula" in subData:
                    subData = self._MakeFormula(subData)

                if "cond_format" in subData:
                    #  special case when we write data but want to exclude it out of applyFormat method
                    self._CondFormat(Formula=subData,i_CellToWrite=location)

                    self._m_Locator.GoToNextColl()
                    self._m_NumOfCellsThatWereWrittenLastTime += 1
                    continue

                if "FormatCellCustom".lower() in subData.lower():
                    self._SetCustumCellFormat(location, subData)
                    continue

                self._XlWrapper.WriteToCell(location, subData)
                self._m_Locator.GoToNextColl()
                self._m_NumOfCellsThatWereWrittenLastTime += 1

    def _ApplyFormat(self):
        """
        The method is passing through the last written cells, and applying the selected format
        as well the method is checking if the value is pass/fail and applyig there a special cell color.
        Notice For NOW THERE is only solid type available
        :param param:
        :return:
        """
        fillType = ""
        fontFormatDict = {'size':None,'color':None}

        cellFormatDict = {'fill_type':fillType,'start_color':None, 'end_color':None}
        startLocation = self._m_LastStartingLocationOfWritting
        self._m_Locator.SetNewLocation(startLocation)
        numOfCells = self._m_NumOfCellsThatWereWrittenLastTime
        for i in range(0, numOfCells):
            cellColor = None
            activeLocation = self._m_Locator.GetCurrLocaion('str')

            #--- CHECK IF PASS OR FAIL OR NA IN CELL
            valueInCell = self._XlWrapper.GetValueFromCell(str(activeLocation))
            valueInCell = str(valueInCell.value)
            fillType = 'solid'
            if activeLocation in self._CellWithFormating:
                self._m_Locator.GoToNextColl()
                continue

            if 'pass' in str(valueInCell).lower() or 'fail' in str(valueInCell).lower() or 'na' in str(valueInCell).lower():
                 cellColor = self._GetPassFailFormat(valueInCell)
            else:
                cellColor = self._m_Formater.GetCellColor()
                if str(self._m_Formater._m_CellColor).lower() == 'white':
                    fillType = 'none'

            cellFormatDict['fill_type'] = fillType
            cellFormatDict['start_color'] = cellColor
            cellFormatDict['end_color'] = cellColor

            fontFormatDict['size'] = self._m_Formater.GetCellFontSize()
            fontFormatDict['color'] = self._m_Formater.GetCellFontColor()

            self._XlWrapper.ApplyFormat(activeLocation, fontFormatDict, cellFormatDict)

            self._m_Locator.GoToNextColl()

        self._CellWithFormating = []

    def _GetPassFailFormat(self, i_Value):
        activeLocation = self._m_Locator.GetCurrLocaion('str')
        cellColor = ''
        if 'pass' in str(i_Value).lower():
            cellColor = self._m_Formater.GetSpecificColor('pass')
        elif 'fail' in str(i_Value).lower():
            cellColor = self._m_Formater.GetSpecificColor('fail')
        elif 'na' in str(i_Value).lower():
            cellColor = self._m_Formater.GetSpecificColor('na')

        return cellColor

    def _SaveNextLocationPerSheet(self):
        """
        The method is savig the location of where data will
        be written next time on current sheet
        :return:
        """
        nameOfSheet =  self._XlWrapper.GetActiveSheetName()
        self._m_LastLocationOnSheet[nameOfSheet] = self._m_Locator.GetCurrLocaion('str')

    def _SetCellWithFormula(self, i_SubData):
        """
        The method is replacing the (curr_Line) to the
        value of currnet line,
        and writing the formula to the current cell
        :param i_SubData:
        :return:
        """
        print(i_SubData)

    def _MakeFormula(self, i_Data):
        """
        The method is rewriting the data to make it
        by replaccing the "(line_Curr)" to the value of current lihe
        example:
        formula=(A(line_Curr)+B(line_Curr))
        "=(A5+B5)"
        :param subData:
        :return: formula to write
        """
        locationAsDic = self._m_Locator.GetCurrLocaion('dic')
        cellValue = i_Data.replace("(line_Curr)", "{currLine}".format(currLine=locationAsDic['Row']))
        formula = "=({form})".format(form=(cellValue.partition("=")[2]))
        formula = formula.replace("|",",")
        return formula

    def _SetCustumCellFormat(self, i_Location, i_SubData):
        """
        The method is setting the decimal place format to the location that was given,
        the decamila place is passed as param
        :param i_Location: location on Excel sheet to set the format
        :param i_SubData:  the precision
        :return:
        """
        numOfPrecision = (str(i_SubData).split("="))[-1]
        strOfZeros = ""
        cell = self._XlWrapper._ws[i_Location]
        cell.number_format = numOfPrecision

    def _CondFormat(self, Formula, i_CellToWrite):
        FormulaParsed = self._MakeFormula(str(Formula)) # send only the right side of the = sign
        FormulaToWrite = FormulaParsed.partition("=")[2]
        self._XlWrapper.ConditionalFormat(FormulaToWrite, i_CellToWrite)
        self._CellWithFormating.append(i_CellToWrite)

if __name__ == "__main__":
    """
    DOCUMENTATION FOR EXCEL FILE WRITER:
    The class allows you wo write the data in deffernet formats:
    - fill the cell with different colors
    - set font size
    - set start location to write the data
    - write formulas to cells (currently only for current line)
    - add more sheets to current test
    Please notice the the class is making save each time when the LINE is written.

    Formatting:

    * startLocation=A2 - the first cell that will be written is A2 and then continue to  same line
    * fontSize=32 - limited to 48 - all next cells will be written with same size till not changed
    * cellColor=jellyfish - sets all next cells with the same color
    list of colors:
    {'black','blue','green','yellow' ,
                            'red' ,"aquamarine",'jellyfish','cream' ,
                            'white' , 'pass', 'fail'}
    NOTICE: pass, fail are green and red colors that suplied by the class,
                                please don't use it manualy!!!

    * SetActiveWorkSheet('SecondTest') - add a new sheet and make it active / make active sheet if sheet exist
        if sheet exist then the next location that the data will be written is the last location were you wrote
        last time
    * formula=(A(line_Curr)) - write formula to the current cell
    * SetDefaultFormat - return to write in default format - as was given to the constructor of the class
    * SetFileNameToSaveResult - set a file to save the data - DON'T FORGET TO WRITE WITH .XLSX
    * FormatCellCustom - set a decimal place atfer the dot (example: 'FormatCellCustom=2','1.2345' will be saved as 1.2345 but shown to user as 1.23 )
    """

    test = ExcelFileWriter()
    fileName = 'Erez.xlsx'
    sheetName = "FirstTest"
    test.SetFileNameToSaveResult(fileName)
    test.SetActiveWorkSheet(sheetName)
    listPassFail = ['pass','fail']
    list_data=['cellColor=jellyfish','13','15','fontSize=20','FormatCellCustom=2','33.1234','fail','asd','pass','qwe', 'formula=(A(line_Curr)*B(line_Curr))']
    # data1 = "fontSize=32 ,a,b,c,d,e"
    test1= ExcelFileWriter('cellColor=aquamarine', 'fontSize=16')

    list_data2 = ['1.16',"3",'pass','fail','PASS',"cond_format=(A(line_Curr) < B(line_Curr))"]
    list_data3= ['1','2','1','2','1','2','1','2','1',"cond_format=(I(line_Curr) >= H(line_Curr))"]
    list_data4 = [ 'fail,cellColor=cream,1G,1200,800,1,32000,2,64,FormatCellCustom=0,formula=(None)/1000,FormatCellCustom=0,formula=(225387+228987)/1000,cond_format=(I(line_Curr) >= H(line_Curr)),FormatCellCustom=0,formula=IF(G(line_Curr)="IMIX"|""|I(line_Curr)*(G(line_Curr)+20)*8/1000),FormatCellCustom=0,25.01,FormatCellCustom=0,100.0,FormatCellCustom=0,0.0,FormatCellCustom=0,0.0,FormatCellCustom=0,0.0' ]
    list_data5 = ['cellColor=cream', '1G', '1200', '800', '1', '32000', '2', '64', 'FormatCellCustom=1',
                  'formula=(None)/1000', 'FormatCellCustom=0', 'formula=(225387+228987)/1000',
                  'cond_format=(I(line_Curr) >= H(line_Curr))', 'FormatCellCustom=0',
                  'formula=IF(G(line_Curr)="IMIX"|""|I(line_Curr)*(G(line_Curr)+20)*8/1000)', 'FormatCellCustom=0', '25.01',
                  'FormatCellCustom=0', '100.0', 'FormatCellCustom=0', '0.0', 'FormatCellCustom=0', '0.0', 'FormatCellCustom=0', '0.0']
    list_data6 = ['cellColor=cream', '1G', '1200', '800', '1', '32000', '2', '64', 'FormatCellCustom=1',
                  'formula=(None)/1000', 'FormatCellCustom=0', 'formula=(225387+228987)/1000',
                  'cond_format=(I(line_Curr) >= H(line_Curr))', 'FormatCellCustom=0',
                  'formula=IF(G(line_Curr)="IMIX"|""|I(line_Curr)*(G(line_Curr)+20)*8/1000)', 'FormatCellCustom=0', '25.01',
                  'FormatCellCustom=0', '100.0', 'FormatCellCustom=0', '0.0', 'FormatCellCustom=0', '0.0', 'FormatCellCustom=0.0%', '0.0']

    # data3 = "fail pass This Is Third Line Of Check"
    # data4 = "startLocation=A4 This Is fail pass fontSize=32 Fourth Check"
    # test.WriteLineWithFormating(test1)
    test.WriteLineWithFormating(list_data6)

    # test.GotoNextLineOrNewLocation()
    # test.SetDefaultFormat()
    # test.WriteLineWithFormating(list_data)

    # test.WriteLineWithFormating(list_data2)

    # test.WriteLineWithFormating(list_data2)
    # test.GotoNextLineOrNewLocation()
    # test.WriteLineWithFormating(data1)
    # test.WriteLineWithFormating(data3)
    # test.SetActiveWorkSheet('SecondTest')
    # test.WriteLineWithFormating(data4)
    # append to the exist file check
    # test1.SetFileNameToSaveResult(fileName)
    # test1.SetActiveWorkSheet('SecondTest')
    # test1.WriteLineWithFormating(list_data)
    # test1.SetDefaultFormat()
    # test1.WriteLineWithFormating(data3)
    # # stop
    # test.WriteLineWithFormating("startLocation=A6 cellColor=jellyfish fontSize=8 FIRST This pass/fail Is First Fail  cellColor=red fontSize=16 You Will paSs Never FAiL Die")
    # some.WriteLineWithFormating(
    #     "startLocation=C3 cellColor=yellow fontSize=12  SECOND This Is Second   cellColor=red fontSize=16 Pass Line Fail To Check")
    # some.GotoNextLineOrNewLocation("A15")
    # some.WriteLineWithFormating("THIRD After C3")
    # test.SetDefaultFormat()
    # test.GotoNextLineOrNewLocation()
    # test.WriteLineWithFormating(
    #     "fontSize=8 FOURTH oK pass/fail fontSize=16 You Will paSs Never FAiL Die")
    # some.SetActiveWorkSheet("SecondTest")
    # some.SetDefaultFormat()
    # some.WriteLineWithFormating(
    #     "FIFTH This is Fail Second  cellColor=red fontSize=16 Pass Will Fail Never Test No 2")
    # some.GotoNextLineOrNewLocation()
    # some.SetActiveWorkSheet("FirstTest")
    # some.WriteLineWithFormating(
    #     "cellColor=cream fontSize=30 SIXTH This is Last Time  cellColor=yellow fontSize=16 Pass I  Fail Never Test Wrote 2 Here")
    # some.SetActiveWorkSheet("SecondTest")
    # some.WriteLineWithFormating(
    #     "cellColor=yellow fontSize=12 This Is Second   cellColor=red fontSize=16 Pass Line Fail To Check")
