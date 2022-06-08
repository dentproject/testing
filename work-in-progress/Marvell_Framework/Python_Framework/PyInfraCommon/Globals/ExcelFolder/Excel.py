###################################################################################
#	Marvell GPL License
#	
#	If you received this File from Marvell, you may opt to use, redistribute and/or
#	modify this File in accordance with the terms and conditions of the General
#	Public License Version 2, June 1991 (the "GPL License"), a copy of which is
#	available along with the File in the license.txt file or by writing to the Free
#	Software Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA 02111-1307 or
#	on the worldwide web at http://www.gnu.org/licenses/gpl.txt.
#	
#	THE FILE IS DISTRIBUTED AS-IS, WITHOUT WARRANTY OF ANY KIND, AND THE IMPLIED
#	WARRANTIES OF MERCHANTABILITY OR FITNESS FOR A PARTICULAR PURPOSE ARE EXPRESSLY
#	DISCLAIMED.  The GPL License provides additional details about this warranty
#	disclaimer.
###################################################################################

from __future__ import print_function
from builtins import chr
from builtins import str
from builtins import range
from builtins import object
__author__ = 'Erez Ashkenazi'
import openpyxl, re

from openpyxl.utils.cell import column_index_from_string

class Excel(object):

    def __init__(self):
        """
        Constructor
        """
        pass

    def set_active_workbook(self, file_location):
        """
        """
        self.WB=openpyxl.load_workbook(file_location, data_only=True) #data_only=true wil show us the value and not the function inside
        return self.WB

    def set_active_worksheet(self, sheet_name):
        """
        Change working sheet to the name inputted
        """
        self.WS = self.WB[str(sheet_name)]
        return self.WS

    def get_cell_value(self,row,col=None, i_isReturnAbsoluteLocation = False):
        """
        Rerutn cell data based on coordination or location.
        if only row is Entered row=B3 than we'll use the info as coordination(B3) and return the info from the cell.
        if row + col is entered we'll search for data based on location
        Ex: string/location Coordination:
        row=Hello / 3
        col=Test1 / B

               A          B
             -----------------------
         1   -           Test1
         2   -
         3   -Hello      result
         4   -

        if i_isReturnAbsoluteLocation is true then the function
        returns as well cell absolute location - B3 (for examlpe above)
        """
        # Todo Need to add the relative culomn AA and so on all subset modules

        if not col: #probably looking for cell Coordination ex: B3
            (row,col)=openpyxl.utils.coordinate_to_tuple(row)

        if type(row) is str:
            result1=self._find_by_name(row)
            (result1_row,result1_column)=openpyxl.utils.coordinate_to_tuple(result1)
        else:
            result1_row=row
            result1_column=1

        if type(col) is str:
            result2=self._find_by_name(col)
            (result2_row,result2_column)=openpyxl.utils.coordinate_to_tuple(result2)
        else:
            result2_row=1
            result2_column=col

        cell_row=max (result1_row,result2_row)
        cell_column=max (result1_column,result2_column)
        cellFullName = str(chr(ord('A') + int(cell_column - 1))) + str(cell_row)

        if i_isReturnAbsoluteLocation:
            answer = self.WS.cell(row=cell_row,column=cell_column).value, cellFullName
        else:
            answer = self.WS.cell(row=cell_row, column=cell_column).value

        return str(answer)

    def active_sheet(self):
        """
        Show the current active sheet
        """
        return self.WB.active

    def sheets_in_workbook(self):
        """
        name all the sheets in the current workbook
        :return:
        """
        return self.WB.get_sheet_names()

    def matrix_dimensions(self):
        """
        Size of the matrix (row,col)
        :return:
        """
        size=self.WS.dimensions
        size=size.split(':')
        return size

    def get_data_by_column(self, col_data=None, remove_values_at_start=0, value=0):
        """
        Get the content of the specific column in the worksheet
        :param value: Return the value=0 of the cell or all the metadata=1 about it
        :param col_data: can be specific col(A,B,C etc) or the data Header of the col like:Device_Id
        :param remove_values_at_start: choice how many values to remove from the start before returning the data
        :return: Vector with all the data (w or w/o the "header")
        """

        if type(col_data)==str:
            col_data=str(col_data)
        if type(col_data)==int:
            return exit('get_data_by_column: Please supply only string: column name or Coordination A,B,C')

        active_column=self._user_input_Coordination_Or_Word(col_data)
        col_range=self.WS[active_column]
        Range_value=self._extract_value_from_range(col_range)
        if value==0:
            Range_value=col_range
        return Range_value[remove_values_at_start:]

    def get_data_by_row(self, row, remove_values_at_start=0,value=0,remove_vales_at_end=None):
        """
        Get the content of the specific row in the worksheet
        :param value: Return the value of the cell or all the metadata about it
        :param row: can be a number of the row(1,2,3,4 etc) or the first data "Header" of the row like:Device_Id
        :param remove_values_at_start: choice how many values to remove from the start before returning the data
        :param remove_vales_at_end: choice how many values to remove from the end before returning the data
        :return: Vector with all the data (w or w/o the "header")
        """
        if type(row)==str:
            row=str(row)
        if type(row) is str:
            cell_cord=self._find_by_name(row)
            active_row=re.search(r'[0-9]+',cell_cord)
            active_row=active_row.group(0)
        elif type(row) is int:
            active_row=str(row)
        row_range=self.WS[active_row]
        Range_value=self._extract_value_from_range(row_range)
        if value==0:
            Range_value=row_range

        return Range_value[remove_values_at_start:remove_vales_at_end]

    def filter_data_in_column(self,str_in_column=None, Data_To_Filter=''):
        """
        Enter Header of a table and get only the Data_To_Filter that is inside
        :param str_in_column: table header name
        :param Data_To_Filter: data to filter
        :return: List with only the cell info containing the data you have asked for
        """
        Data_list=list(self.get_data_by_column(col_data=str_in_column))
        Data_list_tmp=[]
        for cell in Data_list:
            if str(cell.value)==Data_To_Filter:
                Data_list_tmp.append(cell)

        return Data_list_tmp

    def find_all_ovelapping_rows(self,**kwargs):
        """
        search for the all rows that has all those inputs, and return only those rows number
        We should provide as input: row header = value inside
        EX:
        Cores='2',Dir='Uni',CPU='1600',DDR='800',flows='2'

        :param kwargs: multiple argument to provide specific row location: Cores='2',Dir='Uni',CPU='1600',DDR='800',flows='2'
        :return: Set of all overlapping rows
        """

        Dict_of_input_vs_Coordination={}
        input_arg_num=0
        for key, value in kwargs.items():
            Dict_of_input_vs_Coordination[input_arg_num]=self.filter_data_in_column(str_in_column=key,Data_To_Filter=value)
            input_arg_num += 1

        Dict_of_rows_of_All_Args={}
        #get the list of all rows in this column
        for arg in Dict_of_input_vs_Coordination:
            if len(Dict_of_input_vs_Coordination[arg])==0:
                Dict_of_rows_of_All_Args[arg] = Dict_of_input_vs_Coordination[arg]
                continue
            for cell in Dict_of_input_vs_Coordination[arg]:
                if arg not in Dict_of_rows_of_All_Args:
                    Dict_of_rows_of_All_Args[arg]={}
                if type(Dict_of_rows_of_All_Args[arg]) != list:
                    Dict_of_rows_of_All_Args[arg] = []
                Dict_of_rows_of_All_Args[arg].append(cell.row)

        output=[]
        for num in range(0,Dict_of_rows_of_All_Args.__len__()-1):
            if num == 0:
                output=set(Dict_of_rows_of_All_Args[num]).intersection(Dict_of_rows_of_All_Args[num+1])
            else:
                output = output.intersection(Dict_of_rows_of_All_Args[num + 1])

        return output

    def _extract_value_from_range(self,range):
        """
        internal function to extract the value from a range of cells into a vector
        :param range:
        :return:
        """
        Data=[]
        for cell in range:
            if type(cell.value)==int:
                cell.value=int(cell.value)
            Data.append(cell.value)

        # remove empty values from the end
        while Data[-1]==None:
            Data.pop(-1)
            if len(Data)==0:
                break

        return Data

    def _find_by_name(self,input,all_repeats=0):
        """
        Internal Usage find cell by data inside, basically we are looking for headers in the table, and headers should be unique.
        :param input: value we are looking for
        :param all_repeats: if set than all repeats of the same value will be returned with their data sets{B:[<Cell L2 FWD (KS).B34>],E:[<Cell L2 FWD (KS).E26>] }.
                if not set will return the first value found
        """
        rows=self.WS.rows
        tmp_dict={}
        for row in rows:
            for cell in row:
                if str(cell.value).strip()==input.strip():
                    if all_repeats==0:
                        return cell.column + str(cell.row)
                    else:
                        if cell.column not in tmp_dict:
                            tmp_dict[cell.column]=[]
                        tmp_dict[cell.column].append(cell)

        return tmp_dict

    def _user_input_Coordination_Or_Word(self, data):
        """
        Did the user gave us a cell coordination or cell string content
        mostly will be used when searching within columns for column coordination or table header
        :return: the column number
        """
        sheet_max_col=int(re.search(r'[0-9]+',self.matrix_dimensions()[1]).group())

        try:
            inpt_coord=int(column_index_from_string(data))
        except ValueError:
            inpt_coord=sheet_max_col+100

        if inpt_coord<=sheet_max_col:
            active_column=data
        else:
            active_column=self._find_by_name(data)
            active_column=re.search(r'^[a-zA-Z]+',active_column).group(0)

        return active_column

if __name__ == '__main__':
    #from ExcelFolder import Excel
    ex=Excel()
    # print ex
    # print 'Next'
    #w = ex.workbook(file_location=r'\\fileril103\dumps\erezas\Test_Run_Params.xlsx')
    #print w
    # WB=ex.set_active_workbook(file_location='/home/erezas/trash/mmp_automation/Test_Files/Test_Run_Params_CDAL.xlsx')
    WB=ex.set_active_workbook(file_location='/media/local/users/erezas/Automation-devel/Test_Files/Test_Iteration_Benchmark.xlsx')
    # WB=ex.set_active_workbook(file_location='/home/erezas/trash/mmp_automation/Test_Files/temp.xlsx')
    # WS=ex.set_active_worksheet('Environment_Data')
    # WS=ex.set_active_worksheet(r'L2 FWD (KS)')
    WS=ex.set_active_worksheet('KS-L2-FWD')
    #print 'Result is:' ,ex.get_cell_value('VoIP100010',2)
    #b=ex.get_data_by_column('Run',remove_values_at_start=1,value=1)
    # b=ex.get_coordination_by_matches_multi_row_inputs(Board='DB-88F8040-MODULAR',gerrit='1')#Cores='2',Dir='Uni',CPU='1600',DDR='800',flows='2')
    b = ex.find_all_ovelapping_rows(Board='DB-88F8040-MODULAR',gerrit='1',nightly='1')
    print(b)
    exit(1)
    a=ex.get_data_by_row(1, 0,1)
    #print b
    print(a)
    # WB = ex.set_active_workbook(file_location="C:\GITAlexZemtzov\mmp_automation\ExcelFolder\CheckFile.xlsx")
    # WS = ex.set_active_worksheet('Sheet1')
    # Cell = ex.get_cell_value('row5','colD')
    # print Cell
    # Cell = ex.get_cell_value('row5','colD', True)
    # print Cell
